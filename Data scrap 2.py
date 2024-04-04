#pip install (certifi, beautifulsoup4, requests, openpyxl pandas)
#this will generate a excel file within the folder for you to view.
import os
import requests
from bs4 import BeautifulSoup
import certifi
import openpyxl
import pandas as pd

#This is the most important part of the code. This is the URL of the website that we want to scrape.
base_url = "https://www.seek.com.au/software-engineer-jobs/in-Gordon-NSW-2072?distance=25&salaryrange=60000-120000&salarytype=annual&sortmode=ListedDate"

def scrape_seek_jobs(url):
    #this certificate will expire every 90 days, there are ways to get a better authentication method.
    # However I really struggled to get this to work, will update when its solved.
    response = requests.get(url, verify='./Certificate/seek1.cer')
    response.raise_for_status()

    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the elements containing relevant job details, this can easily be done by inspecting the html of the website
    job_cards = soup.find_all('article', {'data-automation': 'normalJob'}) 

    jobs = []
    for card in job_cards: 
        try: 
             
            job_title = card.find('a', {'data-automation': 'jobTitle'}).text.strip()

            # Filter for jobs that have "software engineer" in the title
            if "software engineer".lower() in job_title.lower(): 
                job_title = card.find('a', {'data-automation': 'jobTitle'}).text.strip()
                company_name = card.find('a', {'data-automation': 'jobCompany'}).text.strip()
                location = card.find('a', {'data-automation': 'jobLocation'}).text.strip()  
                age = card.find('span', {'data-automation': 'jobListingDate'}).text.strip() 


                job = {
                'title': job_title,
                'company': company_name,  
                'location': location,
                'age': age 
                 }
                jobs.append(job)
                
                jobs.append(job)

        except Exception as e:
            print(f"An error occurred within a job card: {e}") 

    return jobs


if os.path.exists('seek_jobs.xlsx'):
        print("The file 'seek_jobs.xlsx' already exists. Please delete it and rerun the script.")
        exit()
        
if __name__ == '__main__':
    jobs_data = scrape_seek_jobs(base_url)
    df = pd.DataFrame(jobs_data)
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = 'Title'
    sheet.cell(row=1, column=2).value = 'Company'
    sheet.cell(row=1, column=3).value = 'Location'
    sheet.cell(row=1, column=4).value = 'Age'
    
    for index, row in df.iterrows():
        sheet.cell(row=index + 2, column=1).value = row['title']
        sheet.cell(row=index + 2, column=2).value = row['company']
        sheet.cell(row=index + 2, column=3).value = row['location']
        sheet.cell(row=index + 2, column=4).value = row['age']

workbook.save('seek_jobs.xlsx') 
    